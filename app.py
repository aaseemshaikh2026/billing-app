from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from models import db, User, Client, Role, Employee, Holiday, Leave
from utils import get_date_range_billing, calculate_billing
from datetime import datetime
from calendar import monthrange
import io, os, random, string, re, smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# In-memory OTP store: {email: {'otp': '123456', 'expires': datetime, 'data': {...}}}
pending_registrations = {}

# Email config - set these as environment variables for production
SMTP_EMAIL = os.environ.get('SMTP_EMAIL', '')  # e.g. your Gmail
SMTP_PASSWORD = os.environ.get('SMTP_PASSWORD', '')  # Gmail App Password
SMTP_SERVER = os.environ.get('SMTP_SERVER', 'smtp.gmail.com')
SMTP_PORT = int(os.environ.get('SMTP_PORT', '587'))

def send_otp_email(to_email, otp):
    """Send OTP via email. Falls back to console if SMTP not configured."""
    if not SMTP_EMAIL or not SMTP_PASSWORD:
        print(f"[OTP] Email: {to_email} | OTP: {otp}", flush=True)
        return False
    
    try:
        msg = MIMEMultipart()
        msg['From'] = SMTP_EMAIL
        msg['To'] = to_email
        msg['Subject'] = 'Billing App - Email Verification Code'
        body = f"""<html><body style="font-family:Arial,sans-serif;">
        <div style="max-width:400px;margin:0 auto;padding:30px;border:1px solid #ddd;border-radius:10px;">
            <h2 style="color:#366092;text-align:center;">\U0001f4bc Billing System</h2>
            <p>Your verification code is:</p>
            <div style="text-align:center;padding:20px;background:#f0f4f8;border-radius:8px;margin:15px 0;">
                <span style="font-size:32px;font-weight:bold;letter-spacing:8px;color:#366092;">{otp}</span>
            </div>
            <p style="color:#666;font-size:0.9em;">This code expires in 10 minutes. Do not share it with anyone.</p>
        </div></body></html>"""
        msg.attach(MIMEText(body, 'html'))
        
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(SMTP_EMAIL, SMTP_PASSWORD)
            server.send_message(msg)
        return True
    except Exception as e:
        print(f"[EMAIL ERROR] {e}", flush=True)
        print(f"[OTP FALLBACK] Email: {to_email} | OTP: {otp}", flush=True)
        return False


app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')

db.init_app(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

@login_manager.unauthorized_handler
def unauthorized():
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Authentication required'}), 401
    return redirect(url_for('login'))

with app.app_context():
    db.create_all()

# Helper: verify client belongs to current user
def get_user_client(client_id):
    return Client.query.filter_by(id=client_id, user_id=current_user.id).first()

# ─── AUTH ROUTES ───────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        data = request.json
        user = User.query.filter_by(username=data.get('username')).first()
        if user and user.check_password(data.get('password')) and user.is_active:
            login_user(user)
            return jsonify({'success': True, 'user': {'full_name': user.full_name}})
        return jsonify({'success': False, 'error': 'Invalid credentials'}), 401
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        data = request.json
        for field in ['username', 'password', 'full_name', 'email']:
            if not data.get(field, '').strip():
                return jsonify({'error': f'{field} is required'}), 400
        pw = data['password']
        pw_err = validate_password(pw, data['full_name'])
        if pw_err:
            return jsonify({'error': pw_err}), 400
        if User.query.filter_by(username=data['username']).first():
            return jsonify({'error': 'Username already taken'}), 400
        if User.query.filter_by(email=data['email']).first():
            return jsonify({'error': 'Email already registered'}), 400

        otp = ''.join(random.choices(string.digits, k=6))
        pending_registrations[data['email'].strip().lower()] = {
            'otp': otp,
            'expires': datetime.utcnow().replace(second=0) + __import__('datetime').timedelta(minutes=10),
            'data': {
                'username': data['username'].strip(),
                'full_name': data['full_name'].strip(),
                'email': data['email'].strip(),
                'password': data['password']
            }
        }
        email_sent = send_otp_email(data['email'].strip(), otp)
        msg = f'Verification code sent to {data["email"]}.' if email_sent else f'Verification code sent to {data["email"]}. Check server console for OTP (email not configured).'
        return jsonify({'success': False, 'otp_required': True, 'email': data['email'].strip(), 'message': msg})
    return render_template('login.html')

@app.route('/api/verify-otp', methods=['POST'])
def verify_otp():
    data = request.json
    email = data.get('email', '').strip().lower()
    otp = data.get('otp', '').strip()

    pending = pending_registrations.get(email)
    if not pending:
        return jsonify({'error': 'No pending registration for this email. Please register again.'}), 400
    if pending['otp'] != otp:
        return jsonify({'error': 'Invalid verification code'}), 400
    if datetime.utcnow() > pending['expires']:
        del pending_registrations[email]
        return jsonify({'error': 'Code expired. Please register again.'}), 400

    reg = pending['data']
    user = User(username=reg['username'], full_name=reg['full_name'], email=reg['email'], email_verified=True)
    user.set_password(reg['password'])
    db.session.add(user)
    db.session.commit()
    del pending_registrations[email]
    login_user(user)
    return jsonify({'success': True})

@app.route('/api/resend-otp', methods=['POST'])
def resend_otp():
    email = request.json.get('email', '').strip().lower()
    pending = pending_registrations.get(email)
    if not pending:
        return jsonify({'error': 'No pending registration. Please register again.'}), 400
    otp = ''.join(random.choices(string.digits, k=6))
    pending['otp'] = otp
    pending['expires'] = datetime.utcnow().replace(second=0) + __import__('datetime').timedelta(minutes=10)
    email_sent = send_otp_email(email, otp)
    msg = 'New code sent to your email.' if email_sent else 'New code sent. Check server console.'
    return jsonify({'message': msg})

def validate_password(pw, full_name=''):
    if len(pw) < 8:
        return 'Password must be at least 8 characters'
    if not re.search(r'[A-Z]', pw):
        return 'Password must contain at least one uppercase letter'
    if not re.search(r'[a-z]', pw):
        return 'Password must contain at least one lowercase letter'
    if not re.search(r'[0-9]', pw):
        return 'Password must contain at least one number'
    if not re.search(r'[!@#$%^&*(),.?":{}|<>]', pw):
        return 'Password must contain at least one special character (!@#$%^&*)'
    if full_name:
        for part in full_name.strip().split():
            if len(part) >= 2 and part.lower() in pw.lower():
                return f'Password must not contain your name ("{part}")'
    return None

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/api/current-user')
@login_required
def current_user_info():
    return jsonify({'username': current_user.username, 'full_name': current_user.full_name, 'email': current_user.email})

@app.route('/api/change-password', methods=['POST'])
@login_required
def change_password():
    data = request.json
    if not current_user.check_password(data.get('current_password', '')):
        return jsonify({'error': 'Current password is incorrect'}), 400
    pw_err = validate_password(data.get('new_password', ''), current_user.full_name)
    if pw_err:
        return jsonify({'error': pw_err}), 400
    current_user.set_password(data['new_password'])
    db.session.commit()
    return jsonify({'message': 'Password changed successfully'})

@app.route('/api/security-questions', methods=['GET', 'POST'])
@login_required
def manage_security_questions():
    if request.method == 'POST':
        data = request.json
        for f in ['question_1', 'answer_1', 'question_2', 'answer_2']:
            if not data.get(f, '').strip():
                return jsonify({'error': 'All fields are required'}), 400
        current_user.security_question_1 = data['question_1']
        current_user.set_security_answer(1, data['answer_1'])
        current_user.security_question_2 = data['question_2']
        current_user.set_security_answer(2, data['answer_2'])
        db.session.commit()
        return jsonify({'message': 'Security questions saved'})
    return jsonify({
        'has_questions': bool(current_user.security_question_1),
        'question_1': current_user.security_question_1,
        'question_2': current_user.security_question_2
    })

@app.route('/api/forgot-password', methods=['POST'])
def forgot_password():
    user = User.query.filter_by(username=request.json.get('username')).first()
    if not user or not user.security_question_1:
        return jsonify({'error': 'User not found or security questions not set'}), 400
    return jsonify({'question_1': user.security_question_1, 'question_2': user.security_question_2})

@app.route('/api/reset-password', methods=['POST'])
def reset_password():
    data = request.json
    user = User.query.filter_by(username=data.get('username')).first()
    if not user:
        return jsonify({'error': 'User not found'}), 404
    if not user.check_security_answer(1, data.get('answer_1', '')) or not user.check_security_answer(2, data.get('answer_2', '')):
        return jsonify({'error': 'Security answers are incorrect'}), 400
    pw_err = validate_password(data.get('new_password', ''), user.full_name)
    if pw_err:
        return jsonify({'error': pw_err}), 400
    user.set_password(data['new_password'])
    db.session.commit()
    return jsonify({'message': 'Password reset successfully'})

# ─── MAIN PAGE ─────────────────────────────────────────────────

@app.route('/')
@login_required
def index():
    return render_template('index.html')

# ─── CLIENT ROUTES ─────────────────────────────────────────────

@app.route('/api/clients', methods=['GET', 'POST'])
@login_required
def manage_clients():
    if request.method == 'POST':
        data = request.json
        if not data.get('client_name', '').strip():
            return jsonify({'error': 'Client name is required'}), 400
        try:
            client = Client(user_id=current_user.id, client_name=data['client_name'].strip(), description=data.get('description', ''))
            db.session.add(client)
            db.session.commit()
            return jsonify({'id': client.id, 'client_name': client.client_name, 'description': client.description})
        except Exception:
            db.session.rollback()
            return jsonify({'error': 'Failed to create client'}), 400

    clients = Client.query.filter_by(user_id=current_user.id).all()
    return jsonify([{'id': c.id, 'client_name': c.client_name, 'description': c.description} for c in clients])

@app.route('/api/clients/<int:cid>', methods=['PUT', 'DELETE'])
@login_required
def update_client(cid):
    client = get_user_client(cid)
    if not client:
        return jsonify({'error': 'Client not found'}), 404
    if request.method == 'PUT':
        data = request.json
        if data.get('client_name'):
            client.client_name = data['client_name'].strip()
        if 'description' in data:
            client.description = data['description']
        db.session.commit()
        return jsonify({'id': client.id, 'client_name': client.client_name, 'description': client.description})
    db.session.delete(client)
    db.session.commit()
    return jsonify({'message': 'Client deleted'})

# ─── ROLE ROUTES ───────────────────────────────────────────────

@app.route('/api/clients/<int:cid>/roles', methods=['GET', 'POST'])
@login_required
def manage_roles(cid):
    if not get_user_client(cid):
        return jsonify({'error': 'Client not found'}), 404
    if request.method == 'POST':
        data = request.json
        if not data.get('role_name', '').strip():
            return jsonify({'error': 'Role name is required'}), 400
        if not data.get('hourly_rate_usd') or data['hourly_rate_usd'] <= 0:
            return jsonify({'error': 'Hourly rate must be positive'}), 400
        try:
            role = Role(client_id=cid, role_name=data['role_name'].strip(), hourly_rate_usd=data['hourly_rate_usd'])
            db.session.add(role)
            db.session.commit()
            return jsonify({'id': role.id, 'role_name': role.role_name, 'hourly_rate_usd': role.hourly_rate_usd})
        except Exception:
            db.session.rollback()
            return jsonify({'error': 'Failed to create role'}), 400
    roles = Role.query.filter_by(client_id=cid).all()
    return jsonify([{'id': r.id, 'role_name': r.role_name, 'hourly_rate_usd': r.hourly_rate_usd} for r in roles])

@app.route('/api/roles/<int:rid>', methods=['PUT', 'DELETE'])
@login_required
def update_role(rid):
    role = Role.query.get_or_404(rid)
    if not get_user_client(role.client_id):
        return jsonify({'error': 'Not authorized'}), 403
    if request.method == 'PUT':
        data = request.json
        if data.get('role_name'):
            role.role_name = data['role_name'].strip()
        if 'hourly_rate_usd' in data and data['hourly_rate_usd'] > 0:
            role.hourly_rate_usd = data['hourly_rate_usd']
        db.session.commit()
        return jsonify({'id': role.id, 'role_name': role.role_name, 'hourly_rate_usd': role.hourly_rate_usd})
    try:
        db.session.delete(role)
        db.session.commit()
        return jsonify({'message': 'Role deleted'})
    except Exception:
        db.session.rollback()
        return jsonify({'error': 'Cannot delete role with assigned employees'}), 400

# ─── EMPLOYEE ROUTES ───────────────────────────────────────────

@app.route('/api/clients/<int:cid>/employees', methods=['GET', 'POST'])
@login_required
def manage_employees(cid):
    if not get_user_client(cid):
        return jsonify({'error': 'Client not found'}), 404
    if request.method == 'POST':
        data = request.json
        if not data.get('name', '').strip():
            return jsonify({'error': 'Employee name is required'}), 400
        try:
            emp = Employee(client_id=cid, name=data['name'].strip(), role_id=data['role_id'],
                           billing_start_date=datetime.strptime(data['billing_start_date'], '%Y-%m-%d').date(),
                           is_active=data.get('is_active', True))
            db.session.add(emp)
            db.session.commit()
            return jsonify({'id': emp.id, 'name': emp.name, 'role_id': emp.role_id, 'role_name': emp.role.role_name,
                            'hourly_rate': emp.role.hourly_rate_usd, 'billing_start_date': emp.billing_start_date.isoformat(), 'is_active': emp.is_active})
        except Exception:
            db.session.rollback()
            return jsonify({'error': 'Failed to add employee'}), 400
    emps = Employee.query.filter_by(client_id=cid).all()
    return jsonify([{'id': e.id, 'name': e.name, 'role_id': e.role_id, 'role_name': e.role.role_name,
                     'hourly_rate': e.role.hourly_rate_usd, 'billing_start_date': e.billing_start_date.isoformat(), 'is_active': e.is_active} for e in emps])

@app.route('/api/employees/<int:eid>', methods=['PUT', 'DELETE'])
@login_required
def update_employee(eid):
    emp = Employee.query.get_or_404(eid)
    if not get_user_client(emp.client_id):
        return jsonify({'error': 'Not authorized'}), 403
    if request.method == 'PUT':
        data = request.json
        if data.get('name'):
            emp.name = data['name'].strip()
        if 'role_id' in data:
            emp.role_id = data['role_id']
        if 'billing_start_date' in data:
            emp.billing_start_date = datetime.strptime(data['billing_start_date'], '%Y-%m-%d').date()
        if 'is_active' in data:
            emp.is_active = data['is_active']
        db.session.commit()
        return jsonify({'id': emp.id, 'name': emp.name, 'role_name': emp.role.role_name, 'billing_start_date': emp.billing_start_date.isoformat(), 'is_active': emp.is_active})
    db.session.delete(emp)
    db.session.commit()
    return jsonify({'message': 'Employee deleted'})

# ─── HOLIDAY ROUTES ────────────────────────────────────────────

@app.route('/api/clients/<int:cid>/holidays', methods=['GET', 'POST'])
@login_required
def manage_holidays(cid):
    if not get_user_client(cid):
        return jsonify({'error': 'Client not found'}), 404
    if request.method == 'POST':
        data = request.json
        added = 0
        for h in data.get('holidays', []):
            hdate = datetime.strptime(h['date'], '%Y-%m-%d').date()
            if not Holiday.query.filter_by(client_id=cid, holiday_date=hdate).first():
                db.session.add(Holiday(client_id=cid, holiday_date=hdate, description=h.get('description', ''), year=hdate.year))
                added += 1
        db.session.commit()
        return jsonify({'added': added})
    year = request.args.get('year', type=int)
    q = Holiday.query.filter_by(client_id=cid)
    if year:
        q = q.filter_by(year=year)
    holidays = q.order_by(Holiday.holiday_date).all()
    return jsonify([{'id': h.id, 'date': h.holiday_date.isoformat(), 'description': h.description, 'year': h.year} for h in holidays])

@app.route('/api/holidays/<int:hid>', methods=['DELETE'])
@login_required
def delete_holiday(hid):
    h = Holiday.query.get_or_404(hid)
    if not get_user_client(h.client_id):
        return jsonify({'error': 'Not authorized'}), 403
    db.session.delete(h)
    db.session.commit()
    return jsonify({'message': 'Holiday deleted'})

# ─── LEAVE ROUTES ──────────────────────────────────────────────

@app.route('/api/clients/<int:cid>/leaves', methods=['GET', 'POST'])
@login_required
def manage_leaves(cid):
    if not get_user_client(cid):
        return jsonify({'error': 'Client not found'}), 404
    if request.method == 'POST':
        data = request.json
        emp = Employee.query.get(data['employee_id'])
        if not emp or emp.client_id != cid:
            return jsonify({'error': 'Employee not found'}), 404
        ldate = datetime.strptime(data['leave_date'], '%Y-%m-%d').date()
        if ldate.weekday() >= 5:
            return jsonify({'error': 'Cannot apply leave on a weekend (Saturday/Sunday)'}), 400
        if Holiday.query.filter_by(client_id=cid, holiday_date=ldate).first():
            return jsonify({'error': 'Cannot apply leave on a holiday'}), 400
        if Leave.query.filter_by(employee_id=data['employee_id'], leave_date=ldate).first():
            return jsonify({'error': 'Leave already exists for this date'}), 400
        leave = Leave(employee_id=data['employee_id'], leave_date=ldate, is_half_day=data.get('is_half_day', False))
        db.session.add(leave)
        db.session.commit()
        return jsonify({'id': leave.id, 'employee_id': leave.employee_id, 'employee_name': emp.name,
                        'leave_date': leave.leave_date.isoformat(), 'is_half_day': leave.is_half_day})
    emp_id = request.args.get('employee_id', type=int)
    emps = Employee.query.filter_by(client_id=cid).all()
    emp_ids = [e.id for e in emps]
    q = Leave.query.filter(Leave.employee_id.in_(emp_ids))
    if emp_id:
        q = q.filter_by(employee_id=emp_id)
    leaves = q.order_by(Leave.leave_date.desc()).all()
    return jsonify([{'id': l.id, 'employee_id': l.employee_id, 'employee_name': l.employee.name,
                     'leave_date': l.leave_date.isoformat(), 'is_half_day': l.is_half_day} for l in leaves])

@app.route('/api/leaves/<int:lid>', methods=['PUT', 'DELETE'])
@login_required
def update_leave(lid):
    leave = Leave.query.get_or_404(lid)
    if not get_user_client(leave.employee.client_id):
        return jsonify({'error': 'Not authorized'}), 403
    if request.method == 'PUT':
        data = request.json
        if 'leave_date' in data:
            leave.leave_date = datetime.strptime(data['leave_date'], '%Y-%m-%d').date()
        if 'is_half_day' in data:
            leave.is_half_day = data['is_half_day']
        db.session.commit()
        return jsonify({'id': leave.id, 'leave_date': leave.leave_date.isoformat(), 'is_half_day': leave.is_half_day})
    db.session.delete(leave)
    db.session.commit()
    return jsonify({'message': 'Leave deleted'})

# ─── BILLING ROUTES ────────────────────────────────────────────

@app.route('/api/clients/<int:cid>/billing/monthly', methods=['GET'])
@login_required
def get_monthly_billing(cid):
    if not get_user_client(cid):
        return jsonify({'error': 'Client not found'}), 404
    month = request.args.get('month', type=int)
    year = request.args.get('year', type=int)
    if not month or not year:
        return jsonify({'error': 'Month and year required'}), 400
    start_date = datetime(year, month, 1).date()
    end_date = datetime(year, month, monthrange(year, month)[1]).date()
    emps = Employee.query.filter_by(client_id=cid, is_active=True).all()
    results = []
    for emp in emps:
        if emp.billing_start_date > end_date:
            continue
        r = get_date_range_billing(emp.id, start_date, end_date)
        if r:
            results.append({'employee_id': emp.id, 'employee_name': r['employee_name'], 'role': r['role'],
                            'hourly_rate': r['hourly_rate'], 'working_days': r['working_days'], 'billing_amount': r['billing_amount']})
    return jsonify(results)

@app.route('/api/clients/<int:cid>/billing/calculate', methods=['POST'])
@login_required
def calculate_billing_api(cid):
    if not get_user_client(cid):
        return jsonify({'error': 'Client not found'}), 404
    data = request.json
    start_date = datetime.strptime(data['start_date'], '%Y-%m-%d').date()
    end_date = datetime.strptime(data['end_date'], '%Y-%m-%d').date()
    emps = Employee.query.filter_by(client_id=cid, is_active=True).all()
    results = []
    for emp in emps:
        r = get_date_range_billing(emp.id, start_date, end_date)
        if r:
            results.append(r)
    return jsonify(results)

# ─── EXPORT ROUTES ─────────────────────────────────────────────

@app.route('/api/clients/<int:cid>/export/xlsx', methods=['POST'])
@login_required
def export_xlsx(cid):
    client = get_user_client(cid)
    if not client:
        return jsonify({'error': 'Client not found'}), 404
    data = request.json
    start_date = datetime.strptime(data['start_date'], '%Y-%m-%d').date()
    end_date = datetime.strptime(data['end_date'], '%Y-%m-%d').date()

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "Billing Report"
    headers = ['Employee Name', 'Role', 'Hourly Rate (USD)', 'Working Days', 'Total Hours', 'Billing Amount (USD)']
    ws.append(headers)
    hfill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    hfont = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal='center')

    total = 0
    for emp in Employee.query.filter_by(client_id=cid, is_active=True).all():
        r = get_date_range_billing(emp.id, start_date, end_date)
        if r:
            ws.append([r['employee_name'], r['role'], r['hourly_rate'], r['working_days'], r['working_days'] * 8, r['billing_amount']])
            total += r['billing_amount']
    ws.append(['', '', '', '', 'TOTAL', total])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or '')) for c in col) + 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'{client.client_name}_billing_{start_date}_{end_date}.xlsx')

@app.route('/api/clients/<int:cid>/export/pdf', methods=['POST'])
@login_required
def export_pdf(cid):
    client = get_user_client(cid)
    if not client:
        return jsonify({'error': 'Client not found'}), 404
    data = request.json
    start_date = datetime.strptime(data['start_date'], '%Y-%m-%d').date()
    end_date = datetime.strptime(data['end_date'], '%Y-%m-%d').date()

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('T', parent=styles['Heading1'], fontSize=20, textColor=colors.HexColor('#366092'), spaceAfter=20, alignment=1)
    elements = [
        Paragraph(f'Billing Report - {client.client_name}', title_style),
        Paragraph(f'Period: {start_date} to {end_date}', styles['Normal']),
        Spacer(1, 0.3 * inch)
    ]

    tdata = [['Employee', 'Role', 'Rate (USD)', 'Working Days', 'Hours', 'Amount (USD)']]
    total = 0
    for emp in Employee.query.filter_by(client_id=cid, is_active=True).all():
        r = get_date_range_billing(emp.id, start_date, end_date)
        if r:
            tdata.append([r['employee_name'], r['role'], f"${r['hourly_rate']:.2f}", f"{r['working_days']:.1f}",
                          f"{r['working_days'] * 8:.1f}", f"${r['billing_amount']:.2f}"])
            total += r['billing_amount']
    tdata.append(['', '', '', 'TOTAL', '', f"${total:.2f}"])

    t = Table(tdata, colWidths=[2 * inch, 1.2 * inch, 0.9 * inch, 1 * inch, 0.8 * inch, 1.1 * inch])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11), ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey), ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    return send_file(buf, mimetype='application/pdf', as_attachment=True,
                     download_name=f'{client.client_name}_billing_{start_date}_{end_date}.pdf')

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5001))
    app.run(debug=os.environ.get('FLASK_ENV') != 'production', host='0.0.0.0', port=port)
